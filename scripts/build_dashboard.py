#!/usr/bin/env python3
"""
Visual portfolio dashboard from a DhanPath backup.

Sheets:
  Dashboard        - KPIs + Doughnut (asset class) + Pareto (top holdings)
  Sunburst Data    - hierarchical table ready for a one-click Excel Sunburst
  Sub-category     - Pareto by sub-category
  All Holdings     - full detail
"""
import json, sys
from collections import OrderedDict
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

src = sys.argv[1] if len(sys.argv) > 1 else "/Users/pbhat/Downloads/dhanpath-backup-2026-08-12.json"
out = sys.argv[2] if len(sys.argv) > 2 else "/Users/pbhat/Documents/Bhat/finance-os/portfolio-dashboard.xlsx"

with open(src) as f:
    data = json.load(f)
holdings = data["holdings"]

def hclass(h):
    cls = h.get("assetClass")
    if cls: return cls
    st = (h.get("subType") or "").upper()
    return "EPF / NPS / PPF" if st in ("EPF", "PPF", "NPS") else "Unclassified"

total = sum(h.get("value") or 0 for h in holdings)

HEADER = PatternFill("solid", fgColor="0B5394")
HFONT  = Font(color="FFFFFF", bold=True)
TITLE  = Font(bold=True, size=16, color="0B5394")
SUB    = Font(size=10, color="666666")

def put_headers(ws, row, headers):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.fill = HEADER; c.font = HFONT; c.alignment = Alignment(horizontal="center")

def autosize(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

wb = Workbook()

# ── Dashboard ────────────────────────────────────────────────────────────────
dash = wb.active
dash.title = "Dashboard"
dash["A1"] = "Portfolio Visual Dashboard"
dash["A1"].font = TITLE
dash["A2"] = f"Total: ₹{round(total):,} across {len(holdings)} assets"
dash["A2"].font = SUB

# KPI row
cls_total = OrderedDict()
for h in holdings:
    c = hclass(h)
    cls_total.setdefault(c, 0.0)
    cls_total[c] += h.get("value") or 0
top_cls = max(cls_total, key=cls_total.get)
top_cum = 0
for v in sorted((h.get("value") or 0) for h in holdings)[::-1][:20]:
    top_cum += v
kpis = [("Total Assets", f"{len(holdings)}"),
        ("Net Worth", f"₹{round(total):,}"),
        ("Largest Class", f"{top_cls} ({cls_total[top_cls]/total*100:.1f}%)"),
        ("Top 20 Holdings", f"{top_cum/total*100:.1f}% of portfolio")]
for i, (k, v) in enumerate(kpis):
    r = 4
    c = dash.cell(row=r, column=1 + i*2, value=k)
    c.font = Font(size=9, color="666666")
    dash.merge_cells(start_row=r, start_column=1+i*2, end_row=r, end_column=1+i*2)
    c2 = dash.cell(row=r+1, column=1+i*2, value=v)
    c2.font = Font(bold=True, size=13, color="0B5394")
    dash.merge_cells(start_row=r+1, start_column=1+i*2, end_row=r+1, end_column=1+i*2)

# --- Doughnut: asset class ---
dhead = 9
put_headers(dash, dhead, ["Asset Class", "Value (₹)", "%"])
r = dhead + 1
ORDER = ["Equity","Debt","International","EPF / NPS / PPF","Other","Gold","Cash & Savings","Unclassified"]
for cls in ORDER:
    if cls not in cls_total: continue
    dash.cell(row=r, column=1, value=cls)
    dash.cell(row=r, column=2, value=round(cls_total[cls]))
    dash.cell(row=r, column=3, value=round(cls_total[cls]/total*100, 2))
    r += 1
data_ref = Reference(dash, min_col=1, min_row=dhead+1, max_row=r-1)
val_ref  = Reference(dash, min_col=2, min_row=dhead, max_row=r-1)
dough = DoughnutChart()
dough.add_data(val_ref, titles_from_data=True)
dough.set_categories(data_ref)
dough.title = "By Asset Class"
dough.height = 9; dough.width = 14
dash.add_chart(dough, "I2")

# --- Pareto: top 20 holdings (bar) + cumulative % (line) ---
phead = r + 2
put_headers(dash, phead, ["Holding", "Value (₹)", "Cumulative %"])
top = sorted(holdings, key=lambda h: -(h.get("value") or 0))[:20]
cum = 0
r = phead + 1
for h in top:
    v = h.get("value") or 0
    cum += v
    dash.cell(row=r, column=1, value=(h.get("name") or "")[:40])
    dash.cell(row=r, column=2, value=round(v))
    dash.cell(row=r, column=3, value=round(cum/total*100, 1))
    r += 1
bar_ref  = Reference(dash, min_col=2, min_row=phead, max_row=r-1)
cat_ref  = Reference(dash, min_col=1, min_row=phead+1, max_row=r-1)
cum_ref  = Reference(dash, min_col=3, min_row=phead, max_row=r-1)

bar = BarChart()
bar.type = "col"
bar.add_data(bar_ref, titles_from_data=True)
bar.set_categories(cat_ref)
bar.title = "Top 20 Holdings — Pareto"
bar.height = 12; bar.width = 24
bar.x_axis.delete = True  # hide long category labels to keep it clean
bar.legend = None

line = LineChart()
line.add_data(cum_ref, titles_from_data=True)
line.y_axis.axId = 200
line.y_axis.title = "Cumulative %"
line.y_axis.majorGridlines = None
line.marker = {"symbol":"circle","size":5}
bar.y_axis.axId = 100
bar += line
dash.add_chart(bar, "A25")
autosize(dash, [32, 16, 14, 16, 16, 16, 16])

# ── Sunburst Data (Asset Class -> Sub-category) ──────────────────────────────
sb = wb.create_sheet("Sunburst Data")
sb["A1"] = "Sunburst Data — select A2:D{LAST} → Insert → Charts → Sunburst"
sb["A1"].font = Font(bold=True, color="0B5394")
sb["A2"] = "NOTE: Excel's Sunburst reads hierarchy from left-to-right columns. Class → Sub-category → (opt. Holding)."
sb["A2"].font = SUB
hrow = 4
headers = ["Asset Class", "Sub-category", "Value (₹)"]
put_headers(sb, hrow, headers)
subs = OrderedDict()
for h in holdings:
    subs.setdefault((hclass(h), h.get("subType") or "—"), 0.0)
    subs[(hclass(h), h.get("subType") or "—")] += h.get("value") or 0
r = hrow + 1
for (cls, sub), val in subs.items():
    sb.cell(row=r, column=1, value=cls)
    sb.cell(row=r, column=2, value=sub)
    sb.cell(row=r, column=3, value=round(val))
    r += 1
autosize(sb, [22, 30, 16])

# ── Sub-category Pareto ──────────────────────────────────────────────────────
pc = wb.create_sheet("Sub-category Pareto")
pc["A1"] = "Sub-category Concentration (fewer, cleaner bars than per-holding)"
pc["A1"].font = Font(bold=True, color="0B5394")
hrow = 3
put_headers(pc, hrow, ["Sub-category", "Asset Class", "Value (₹)", "Cumulative %"])
rows = sorted(subs.items(), key=lambda kv: -kv[1])
cum = 0
r = hrow + 1
for (cls, sub), val in rows:
    cum += val
    pc.cell(row=r, column=1, value=sub)
    pc.cell(row=r, column=2, value=cls)
    pc.cell(row=r, column=3, value=round(val))
    pc.cell(row=r, column=4, value=round(cum/total*100, 1))
    r += 1
bar_ref = Reference(pc, min_col=3, min_row=hrow, max_row=r-1)
cat_ref = Reference(pc, min_col=1, min_row=hrow+1, max_row=r-1)
cum_ref = Reference(pc, min_col=4, min_row=hrow, max_row=r-1)
bar2 = BarChart(); bar2.type = "col"
bar2.add_data(bar_ref, titles_from_data=True)
bar2.set_categories(cat_ref)
bar2.title = "Pareto by Sub-category"
bar2.height = 14; bar2.width = 28
bar2.x_axis.delete = True
bar2.legend = None
line2 = LineChart()
line2.add_data(cum_ref, titles_from_data=True)
line2.y_axis.axId = 200; line2.y_axis.majorGridlines = None
line2.marker = {"symbol":"circle","size":5}
bar2.y_axis.axId = 100
bar2 += line2
pc.add_chart(bar2, "F3")
autosize(pc, [30, 20, 16, 16])

wb.save(out)
print("Wrote", out)
print("Asset classes:")
for cls, v in cls_total.items():
    print(f"  {cls:22s} ₹{round(v):>12,}  {v/total*100:5.2f}%")
print(f"Top-20 holdings = {top_cum/total*100:.1f}% of portfolio")
