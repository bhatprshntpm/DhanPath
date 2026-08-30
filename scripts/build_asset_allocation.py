#!/usr/bin/env python3
"""
Build a detailed asset-allocation Excel from a DhanPath backup JSON.

Sheets:
  1. Asset Allocation (summary)  - asset class + sub-category rollup with %
  2. All Holdings                - every asset, its category, sub-category,
                                   current value, % of total, and hold type
"""
import json
import sys
from collections import OrderedDict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BACKUP = sys.argv[1] if len(sys.argv) > 1 else "/Users/pbhat/Downloads/dhanpath-backup-2026-08-12.json"
OUT    = sys.argv[2] if len(sys.argv) > 2 else "/Users/pbhat/Documents/Bhat/finance-os/asset-allocation.xlsx"

with open(BACKUP, "r", encoding="utf-8") as f:
    data = json.load(f)

holdings = data["holdings"]
total = sum(h.get("value") or 0 for h in holdings)

# ── Styling helpers ──────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="0B5394")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT  = Font(bold=True, size=16, color="0B5394")
SUB_FONT    = Font(size=10, color="666666")
CAT_FILL    = PatternFill("solid", fgColor="D9E2F3")
CAT_FONT    = Font(bold=True, size=11, color="1F3864")
SUBCAT_FILL = PatternFill("solid", fgColor="F2F5FC")

def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

def inr(v):
    return round(v) if v == int(v) else round(v, 2)

# ── Sheet 1: Asset Allocation summary ────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "Asset Allocation"

ws["A1"] = "DhanPath — Detailed Asset Allocation"
ws["A1"].font = TITLE_FONT
ws["A2"] = f"As of {data['snapshots'][-1]['date'] if data['snapshots'] else ''} · Total investable assets: ₹{inr(total):,}"
ws["A2"].font = SUB_FONT

headers = ["Asset Class", "Sub-Category", "No. of Assets", "Current Value (₹)", "% of Portfolio"]
hrow = 4
for i, h in enumerate(headers, start=1):
    ws.cell(row=hrow, column=i, value=h)
style_header(ws, hrow, len(headers))

# Roll up by assetClass -> subType, and track per-asset entries for sorting.
from collections import defaultdict
sub_map = defaultdict(lambda: {"count": 0, "value": 0.0, "assets": []})
def holding_class(h):
    """Asset class, with EPF/PPF/NPS retirement fallback when assetClass is missing."""
    cls = h.get("assetClass")
    if cls:
        return cls
    st = (h.get("subType") or "").upper()
    if st in ("EPF", "PPF", "NPS"):
        return "EPF / NPS / PPF"
    return "Unclassified"

for h in holdings:
    cls  = holding_class(h)
    sub  = h.get("subType") or "—"
    val  = h.get("value") or 0
    sub_map[(cls, sub)]["count"] += 1
    sub_map[(cls, sub)]["value"] += val
    sub_map[(cls, sub)]["assets"].append((h.get("name", ""), val))

ORDER = ["Equity", "Debt", "International", "Gold", "Cash & Savings", "EPF / NPS / PPF", "Other"]
def sort_key(item):
    cls = item[0][0]
    return (ORDER.index(cls) if cls in ORDER else len(ORDER), item[0][1])

r = hrow + 1
for (cls, sub), info in sorted(sub_map.items(), key=sort_key):
    ws.cell(row=r, column=1, value=cls).font = CAT_FONT
    ws.cell(row=r, column=1).fill = CAT_FILL
    ws.cell(row=r, column=2, value=sub).font = CAT_FONT
    ws.cell(row=r, column=2).fill = CAT_FILL
    ws.cell(row=r, column=3, value=info["count"]).alignment = Alignment(horizontal="center")
    ws.cell(row=r, column=4, value=round(info["value"]))
    ws.cell(row=r, column=5, value=round(info["value"] / total * 100, 2))
    for col in (1, 2, 4, 5):
        ws.cell(row=r, column=col).fill = CAT_FILL
    r += 1

# Asset-class totals row
cls_map = OrderedDict()
for h in holdings:
    cls = holding_class(h)
    cls_map.setdefault(cls, {"count": 0, "value": 0.0})
    cls_map[cls]["count"] += 1
    cls_map[cls]["value"] += h.get("value") or 0

r += 1
ws.cell(row=r, column=1, value="TOTAL").font = Font(bold=True, size=11)
ws.cell(row=r, column=2, value="All categories").font = Font(bold=True, size=11)
ws.cell(row=r, column=3, value=len(holdings)).font = Font(bold=True)
ws.cell(row=r, column=3).alignment = Alignment(horizontal="center")
ws.cell(row=r, column=4, value=round(total)).font = Font(bold=True)
tc = ws.cell(row=r, column=5, value=100.0); tc.font = Font(bold=True)
for col in range(1, 6):
    ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor="FFF2CC")
    ws.cell(row=r, column=col).border = __import__("openpyxl").styles.Border(
        top=__import__("openpyxl").styles.Side(style="thin"), bottom=__import__("openpyxl").styles.Side(style="thin"))

# Per-asset-class percentage breakdown block
r += 2
ws.cell(row=r, column=1, value="By Asset Class").font = Font(bold=True, size=12, color="0B5394")
r += 1
for i, h in enumerate(headers[:4] + ["% of Portfolio"], start=1):
    ws.cell(row=r, column=i, value=h)
style_header(ws, r, 5)
r += 1
for cls, info in cls_map.items():
    ws.cell(row=r, column=1, value=cls).font = CAT_FONT
    ws.cell(row=r, column=3, value=info["count"]).alignment = Alignment(horizontal="center")
    ws.cell(row=r, column=4, value=round(info["value"]))
    ws.cell(row=r, column=5, value=round(info["value"] / total * 100, 2))
    for col in (1, 4, 5):
        ws.cell(row=r, column=col).fill = CAT_FILL
    ws.cell(row=r, column=1).fill = CAT_FILL
    r += 1

autosize(ws, [22, 26, 14, 20, 18])

# ── Sheet 2: All Holdings ────────────────────────────────────────────────────
ws2 = wb.create_sheet("All Holdings")
ws2["A1"] = "DhanPath — All Holdings Detail"
ws2["A1"].font = TITLE_FONT
ws2["A2"] = f"Total: {len(holdings)} assets · ₹{inr(total):,}"
ws2["A2"].font = SUB_FONT

h2 = ["Asset", "Ticker / ISIN", "Type", "Asset Class", "Sub-Category", "Qty", "Current Value (₹)", "% of Portfolio"]
hrow2 = 4
for i, h in enumerate(h2, start=1):
    ws2.cell(row=hrow2, column=i, value=h)
style_header(ws2, hrow2, len(h2))

r = hrow2 + 1
sorted_holdings = sorted(holdings, key=lambda h: (ORDER.index(holding_class(h)) if holding_class(h) in ORDER else len(ORDER), -(h.get("value") or 0)))
for h in sorted_holdings:
    cls = holding_class(h)
    val = h.get("value") or 0
    ws2.cell(row=r, column=1, value=h.get("name", ""))
    ws2.cell(row=r, column=2, value=h.get("ticker", ""))
    ws2.cell(row=r, column=3, value=h.get("type", ""))
    ws2.cell(row=r, column=4, value=cls).font = Font(bold=(cls in ("Equity", "Debt", "International", "Gold")))
    ws2.cell(row=r, column=5, value=h.get("subType", ""))
    ws2.cell(row=r, column=6, value=h.get("qty"))
    ws2.cell(row=r, column=7, value=round(val))
    ws2.cell(row=r, column=8, value=round(val / total * 100, 2))
    for col in (6, 7, 8):
        ws2.cell(row=r, column=col).alignment = Alignment(horizontal="right")
    r += 1

# Totals row
ws2.cell(row=r, column=1, value="TOTAL").font = Font(bold=True)
ws2.cell(row=r, column=7, value=round(total)).font = Font(bold=True)
ws2.cell(row=r, column=8, value=100.0).font = Font(bold=True)
for col in range(1, 9):
    ws2.cell(row=r, column=col).fill = PatternFill("solid", fgColor="FFF2CC")

autosize(ws2, [44, 20, 12, 18, 26, 12, 20, 16])

wb.save(OUT)
print(f"Wrote {OUT}")
print(f"Total assets: {len(holdings)}, total value: ₹{inr(total):,}")
print("\nBy asset class:")
for cls, info in cls_map.items():
    print(f"  {cls:22s} {info['count']:4d} assets  ₹{inr(info['value']):>12,}  {info['value']/total*100:5.2f}%")
